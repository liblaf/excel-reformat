from ast import pattern
import re
import sys
from tkinter import Toplevel

import openpyxl
import openpyxl.workbook.workbook
import openpyxl.worksheet.worksheet
import openpyxl.cell.cell


def find(
    sheet: openpyxl.worksheet.worksheet.Worksheet,
    pattern: str = r"部门",
) -> openpyxl.cell.cell.Cell:
    for row in sheet.iter_rows(max_col=1):
        row: tuple[openpyxl.cell.cell.Cell] = row
        for cell in row:
            cell: openpyxl.cell.cell.Cell = cell
            if re.match(pattern=pattern, string=str(cell.value)):
                print(
                    f'found "{cell.value}" at {cell.coordinate} matches pattern "{pattern}"'
                )
                return cell
    return None


def main(
    input_filename: str,
    output_filename: str = "output.xlsx",
    output_sheetname: str = "result",
):
    workbook: openpyxl.workbook.workbook.Workbook = openpyxl.load_workbook(
        filename=input_filename
    )
    sheet: openpyxl.worksheet.worksheet.Worksheet = workbook.active
    print(f"Opening {sheet} ...")
    top_left: openpyxl.cell.cell.Cell = find(sheet=sheet, pattern=r"部门")
    bottom_left: openpyxl.cell.cell.Cell = find(sheet=sheet, pattern=r"总计")
    width: int = 5
    result: openpyxl.worksheet.worksheet.Worksheet = workbook.create_sheet(
        title=output_sheetname
    )
    active_department: str = None
    departments: dict[str, int] = {}
    types: dict[str, int] = {}
    result.cell(row=1, column=1, value="费用")
    for row in sheet.iter_rows(
        min_row=top_left.row + 1,
        max_row=bottom_left.row - 1,
        min_col=top_left.column,
        max_col=top_left.column + width - 1,
    ):
        row: tuple[openpyxl.cell.cell.Cell] = row
        department: str = row[0].value
        type: str = row[1].value
        value = row[-1].value
        if department:
            if department == "(空白)":
                continue
            active_department: str = department
            if department not in departments:
                departments[department] = len(departments) + 2
                result.cell(
                    row=1,
                    column=departments[department],
                    value=department,
                )
        if type:
            if type == "(空白)":
                continue
            if type not in types:
                types[type] = len(types) + 2
                result.cell(row=types[type], column=1, value=type)
            result.cell(
                row=types[type], column=departments[active_department], value=value
            )
            print(f"{active_department} {type}: {row[-1].value}")
    print(f'Save as {result} in "{output_filename}" ...')
    workbook.save(filename=output_filename)


if __name__ == "__main__":
    if len(sys.argv) > 1:
        input_filename = sys.argv[1]
        output_filename = sys.argv[1]
        output_sheetname = "result"
    else:
        input_filename: str = input("Input filename: ")
        output_filename: str = input_filename
        output_filename: str = (
            input(f"Output filename [{output_filename}]: ") or output_filename
        )
        output_sheetname: str = "result"
        output_sheetname: str = (
            input(f"Output Worksheet [{output_sheetname}]: ") or output_sheetname
        )
    main(
        input_filename=input_filename,
        output_filename=output_filename,
        output_sheetname=output_sheetname,
    )
